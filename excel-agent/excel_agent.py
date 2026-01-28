import sys
from pathlib import Path

# Add shared module to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from claude_agent_sdk import ClaudeAgentOptions, ClaudeSDKClient, AssistantMessage, TextBlock, ResultMessage, create_sdk_mcp_server, tool, HookMatcher
import json

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

from shared.cli import AgentCLI, CLIArgument, run_agent_cli

@tool("create_spreadsheet", "Create a new Excel spreadsheet with data", {
    "filename": str,
    "sheet_name": str,
    "data": list,
    "headers": list
})
async def create_spreadsheet(args):
    filename = args["filename"]
    sheet_name = args.get("sheet_name", "Sheet1")
    data = args["data"]
    headers = args.get("headers", [])

    # Ensure .xlsx extension
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add headers if provided
    row_offset = 0
    if headers:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        row_offset = 1

    # Add data rows
    for row_idx, row_data in enumerate(data, start=1 + row_offset):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, (len(headers) or len(data[0]) if data else 0) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filename)

    return {
        "content": [{
            "type": "text",
            "text": json.dumps({
                "status": "success",
                "filename": filename,
                "sheet_name": sheet_name,
                "rows": len(data),
                "columns": len(headers) if headers else (len(data[0]) if data else 0),
            }, indent=2)
        }]
    }

@tool("read_spreadsheet", "Read data from an Excel spreadsheet", {
    "filename": str,
    "sheet_name": str,
    "range": str
})
async def read_spreadsheet(args):
    filename = args["filename"]
    sheet_name = args.get("sheet_name")
    cell_range = args.get("range")

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        # Read with pandas
        df = pd.read_excel(filename, sheet_name=sheet_name or 0)

        # Convert to records
        headers = list(df.columns)
        data = df.to_dict(orient="records")

        # Get sheet names
        wb = load_workbook(filename, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "filename": filename,
                    "sheet_name": sheet_name or sheet_names[0],
                    "available_sheets": sheet_names,
                    "headers": headers,
                    "row_count": len(data),
                    "data": data[:100],  # Limit to 100 rows
                    "truncated": len(data) > 100,
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }

@tool("edit_spreadsheet", "Edit cells in an existing spreadsheet", {
    "filename": str,
    "sheet_name": str,
    "edits": list
})
async def edit_spreadsheet(args):
    filename = args["filename"]
    sheet_name = args.get("sheet_name")
    edits = args["edits"]

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active

        edited_cells = []
        for edit in edits:
            cell = edit["cell"]
            value = edit["value"]
            ws[cell] = value
            edited_cells.append(cell)

        wb.save(filename)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "status": "success",
                    "filename": filename,
                    "edited_cells": edited_cells,
                    "edit_count": len(edited_cells),
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }


@tool("add_formula", "Add Excel formulas to cells", {
    "filename": str,
    "sheet_name": str,
    "formulas": list
})
async def add_formula(args):
    filename = args["filename"]
    sheet_name = args.get("sheet_name")
    formulas = args["formulas"]

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active

        added_formulas = []
        for item in formulas:
            cell = item["cell"]
            formula = item["formula"]
            # Ensure formula starts with =
            if not formula.startswith("="):
                formula = "=" + formula
            ws[cell] = formula
            added_formulas.append({"cell": cell, "formula": formula})

        wb.save(filename)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "status": "success",
                    "filename": filename,
                    "formulas_added": added_formulas,
                    "count": len(added_formulas),
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }


@tool("get_spreadsheet_info", "Get metadata about a spreadsheet", {
    "filename": str
})
async def get_spreadsheet_info(args):
    filename = args["filename"]

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        wb = load_workbook(filename)

        sheets_info = []
        total_formulas = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formula_count = 0
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0

            # Count formulas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1

            total_formulas += formula_count
            sheets_info.append({
                "name": sheet_name,
                "rows": row_count,
                "columns": col_count,
                "formula_count": formula_count,
            })

        wb.close()

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "filename": filename,
                    "sheet_count": len(sheets_info),
                    "total_formulas": total_formulas,
                    "sheets": sheets_info,
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }

@tool("format_cells", "Apply formatting to cells", {
    "filename": str,
    "sheet_name": str,
    "formats": list
})
async def format_cells(args):
    filename = args["filename"]
    sheet_name = args.get("sheet_name")
    formats = args["formats"]

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active

        formatted_ranges = []

        for fmt in formats:
            cell_range = fmt.get("range")
            if not cell_range:
                continue

            # Build style objects
            font_obj = None
            if "font" in fmt:
                font_config = fmt["font"]
                font_obj = Font(
                    bold=font_config.get("bold", False),
                    color=font_config.get("color"),
                    size=font_config.get("size"),
                )

            fill_obj = None
            if "fill" in fmt:
                fill_config = fmt["fill"]
                color = fill_config.get("color", "FFFFFF")
                fill_obj = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid",
                )

            alignment_obj = None
            if "alignment" in fmt:
                align_config = fmt["alignment"]
                alignment_obj = Alignment(
                    horizontal=align_config.get("horizontal"),
                    vertical=align_config.get("vertical"),
                )

            number_format = fmt.get("number_format")

            # Apply to range - handle single cell or range
            if ":" in cell_range:
                cells = ws[cell_range]
                for row in cells:
                    for cell in row:
                        if font_obj:
                            cell.font = font_obj
                        if fill_obj:
                            cell.fill = fill_obj
                        if alignment_obj:
                            cell.alignment = alignment_obj
                        if number_format:
                            cell.number_format = number_format
            else:
                cell = ws[cell_range]
                if font_obj:
                    cell.font = font_obj
                if fill_obj:
                    cell.fill = fill_obj
                if alignment_obj:
                    cell.alignment = alignment_obj
                if number_format:
                    cell.number_format = number_format

            formatted_ranges.append(cell_range)

        wb.save(filename)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "status": "success",
                    "filename": filename,
                    "formatted_ranges": formatted_ranges,
                    "count": len(formatted_ranges),
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }


@tool("evaluate_spreadsheet", "Evaluate spreadsheet quality against best practices", {"filename": str})
async def evaluate_spreadsheet(args: dict) -> dict:
    filename = args["filename"]

    if not Path(filename).exists():
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": f"File not found: {filename}"})
            }]
        }

    try:
        wb = load_workbook(filename)

        score = 0
        issues = []
        suggestions = []

        # Excel error types to check
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']

        total_cells = 0
        formula_cells = 0
        error_cells = []
        empty_data_cells = []
        formatted_headers = 0
        total_headers = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            if max_row == 0 or max_col == 0:
                continue

            # Check header row (row 1)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    total_headers += 1
                    # Check if header is formatted (bold or has fill)
                    if cell.font.bold or (cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000'):
                        formatted_headers += 1

            # Check data cells
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    total_cells += 1

                    # Check for formulas
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_cells += 1

                    # Check for Excel errors (need to load with data_only=True for this)
                    if cell.value and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in str(cell.value):
                                error_cells.append(f"{sheet_name}!{cell.coordinate}: {err}")
                                break

                    # Check for empty cells in data range
                    if cell.value is None:
                        empty_data_cells.append(f"{sheet_name}!{cell.coordinate}")

        wb.close()

        # Also check with data_only to find formula errors
        wb_data = load_workbook(filename, data_only=True)
        for sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in str(cell.value):
                                loc = f"{sheet_name}!{cell.coordinate}: {err}"
                                if loc not in error_cells:
                                    error_cells.append(loc)
        wb_data.close()

        # Scoring
        # 1. Formula errors (30 points) - 0 tolerance
        if len(error_cells) == 0:
            score += 30
        else:
            issues.append(f"Found {len(error_cells)} formula error(s)")
            for err in error_cells[:5]:  # Show first 5
                issues.append(f"  - {err}")
            suggestions.append("Fix all formula errors before finalizing")

        # 2. Formula usage (25 points)
        if total_cells > 0:
            formula_ratio = formula_cells / total_cells
            if formula_ratio >= 0.1:  # At least 10% formulas
                score += 25
            elif formula_ratio >= 0.05:
                score += 15
                suggestions.append("Consider using more formulas instead of hardcoded values")
            else:
                score += 5
                issues.append("Very few formulas used - mostly hardcoded values")
                suggestions.append("Use Excel formulas for calculations to make spreadsheet dynamic")
        else:
            score += 25  # Empty spreadsheet gets full points here

        # 3. Header formatting (25 points)
        if total_headers > 0:
            header_ratio = formatted_headers / total_headers
            if header_ratio >= 0.8:
                score += 25
            elif header_ratio >= 0.5:
                score += 15
                suggestions.append("Format remaining headers with bold or background color")
            else:
                score += 5
                issues.append("Headers are not properly formatted")
                suggestions.append("Apply bold text and/or yellow background to header row")
        else:
            issues.append("No headers found")
            suggestions.append("Add descriptive column headers to row 1")

        # 4. Data completeness (20 points)
        if total_cells > 0:
            empty_ratio = len(empty_data_cells) / total_cells
            if empty_ratio <= 0.05:  # Less than 5% empty
                score += 20
            elif empty_ratio <= 0.15:
                score += 10
                suggestions.append("Some data cells are empty - verify if intentional")
            else:
                issues.append(f"{len(empty_data_cells)} empty cells in data range")
                suggestions.append("Fill in missing data or remove empty rows/columns")
        else:
            score += 20  # Empty spreadsheet

        passing = score >= 75

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "score": score,
                    "max_score": 100,
                    "passing": passing,
                    "breakdown": {
                        "formula_errors": 30 if len(error_cells) == 0 else 0,
                        "formula_usage": min(25, score - (30 if len(error_cells) == 0 else 0)),
                        "header_formatting": formatted_headers,
                        "data_completeness": 20 if total_cells == 0 or len(empty_data_cells) / total_cells <= 0.05 else 0,
                    },
                    "stats": {
                        "total_cells": total_cells,
                        "formula_cells": formula_cells,
                        "error_count": len(error_cells),
                        "empty_cells": len(empty_data_cells),
                        "formatted_headers": formatted_headers,
                        "total_headers": total_headers,
                    },
                    "issues": issues,
                    "suggestions": suggestions,
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({"error": str(e)})
            }]
        }


async def log_excel_operation(input_data, tool_use_id, context):
    tool_name = input_data.get("tool_name")

    if tool_name.startswith("mcp__excel__"):
        operation = tool_name.replace("mcp__excel__", "")
        print(f"[EXCEL-OP] {operation}")
    elif tool_name == "Skill":
        skill_name = input_data.get("tool_input", {}).get("skill", "unknown")
        print(f"[SKILL] Invoking: {skill_name}")

    return {}


excel_tools = create_sdk_mcp_server(
    name="excel",
    version="1.0.0",
    tools=[create_spreadsheet, read_spreadsheet, edit_spreadsheet, add_formula, get_spreadsheet_info, format_cells, evaluate_spreadsheet]
)

async def run_agent(args) -> None:
    """Run the Excel agent."""

    # Build options with conditional hooks
    options = ClaudeAgentOptions(
        mcp_servers={"excel": excel_tools},
        allowed_tools=[
            "mcp__excel__create_spreadsheet",
            "mcp__excel__read_spreadsheet",
            "mcp__excel__edit_spreadsheet",
            "mcp__excel__add_formula",
            "mcp__excel__get_spreadsheet_info",
            "mcp__excel__format_cells",
            "mcp__excel__evaluate_spreadsheet",
            "Skill",
        ],
        cwd=str(Path(__file__).parent),
        setting_sources=["project"],
        hooks={
            "PreToolUse": [HookMatcher(hooks=[log_excel_operation])]
        } if args.verbose else {}
    )

    # Determine prompt based on CLI args
    if args.query:
        prompt = args.query
    elif args.evaluate:
        prompt = f"Evaluate the quality of {args.evaluate} using the evaluate_spreadsheet tool and report any issues with suggestions for improvement."
    elif args.create_budget:
        output = args.output or "budget_tracker.xlsx"
        prompt = f"Create a budget tracker spreadsheet following best practices from the xlsx skill. Include sample data, formulas for totals, and proper formatting. Save to {output}"
    elif args.create_sales:
        output = args.output or "sales_report.xlsx"
        prompt = f"Create a sales report spreadsheet following best practices from the xlsx skill. Include sample data, formulas for totals and averages, and proper formatting. Save to {output}"
    elif args.analyze:
        prompt = f"Read and analyze the spreadsheet at {args.analyze}. Summarize the data structure, key metrics, and any insights."
    else:
        prompt = "List what spreadsheet operations you can help with and describe the xlsx skill best practices."

    async with ClaudeSDKClient(options=options) as client:
        await client.query(prompt)

        async for message in client.receive_response():
            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        print(f"{block.text}")

            elif isinstance(message, ResultMessage):
                if args.verbose:
                    print(f"[RESULT]: {message.result}")


# =============================================================================
# CLI Configuration
# =============================================================================

cli = AgentCLI(
    name="Excel Agent",
    description="AI-powered spreadsheet creation and analysis using Claude Agent SDK",
    arguments=[
        CLIArgument(name="--query", short="-q", help="Direct query to execute"),
        CLIArgument(name="--create-budget", help="Create budget tracker spreadsheet", action="store_true"),
        CLIArgument(name="--create-sales", help="Create sales report spreadsheet", action="store_true"),
        CLIArgument(name="--analyze", short="-a", help="Analyze existing spreadsheet file"),
        CLIArgument(name="--evaluate", short="-e", help="Evaluate spreadsheet quality"),
        CLIArgument(name="--output", short="-o", help="Output filename for created spreadsheets"),
        CLIArgument(name="--verbose", short="-v", help="Show detailed tool usage", action="store_true"),
    ],
    epilog="""
Examples:
  python excel_agent.py -q "Create an employee directory spreadsheet"
  python excel_agent.py --create-sales -o monthly_sales.xlsx
  python excel_agent.py --create-budget -v
  python excel_agent.py --evaluate existing_file.xlsx
  python excel_agent.py --analyze sales_report.xlsx
"""
)


if __name__ == "__main__":
    exit_code = run_agent_cli(cli, run_agent)
    sys.exit(exit_code)
